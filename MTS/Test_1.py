import math
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

# word_1 = input()  # Название первого файла
# word_2 = input()  # Название второго файла

wb = Workbook('Книга1.xlsx')
ws = wb.active # доступ к созданному по умолчанию листу
name_sheet = wb.sheetnames # список рабочих листов
print('wb',wb)
print('ws',ws.cell(row=1, column=1))



# wb = Workbook() # создаем объект книги
# ws = wb.active # доступ к созданному по умолчанию листу
# ws.append(["Имена первого Объекта", "Координаты первого объекта",
#            "Координаты ближайшего второго объекта", "Расстояние"])
#
# number = 100
# i=2
# amount=0
#
# while (i < number + 2):
#     txt = 0
#     txt = txt.replace("\\left(", "")
#     txt = txt.replace("\\right)", "")
#     ws.cell(row=i, column=3).value = txt
#     ws.cell(row=i, column=5).value = f'{}'
#     ws.cell(row=i, column=6).value = f'{}'
#     ws.cell(row=i, column=7).value = f'{}'
#     ws.cell(row=i, column=8).value = f'{}'
#     ws.cell(row=i, column=9).value = f'{}'
#     i+=1
#
# wb.save("Результаты.xlsx") # Сохраняем