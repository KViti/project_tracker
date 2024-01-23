from sympy import *
import random
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2", "Параметр 3" ])
def TRU(a, b, c):  # Функция для генерации нового набора значений
    a, b, c = random.randint(400, 600)/10, random.randint(100, 300)/10, random.randint(150, 400)/10
    return a, b, c

for i in range(2, 101, 2):
    a, b, c = random.randint(400, 600)/10, random.randint(100, 300)/10, random.randint(150, 400)/10
    K = round((b-(100-a)*c/100)/a, 5)  # Ответ
    while (round(K * 100 - int(K * 100), 5) != 0 or K<=0 or K>=1 or a==b or a==c or b==c or round(K,2)==0):  # Пока после запятой более 1 цифры
        a, b, c = TRU(a, b, c)  # создаём новый набор
        K = (b-(100-a)*c/100)/a  # Ответ выражения меняем на новое
    txt = 'В городе ' + str(round(a,5)) + '% взрослого населения— мужчины. Пенсионеры составляют ' + str(round(b,5)) + '% взрослого населения, причём доля пенсионеров среди женщин равна ' + str(round(c,5)) + '%. Для социологического опроса выбран случайным образом мужчина, проживающий в этом городе. Найдите вероятность события «выбранный мужчина является пенсионером».'
    txt = txt.replace('.0%', '%')
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = K
    ws.cell(row=i, column=7).value = a
    ws.cell(row=i, column=8).value = b
    ws.cell(row=i, column=9).value = c

    a, b, c = random.randint(400, 600)/10, random.randint(100, 300)/10, random.randint(150, 400)/10
    K = round((b-(100-a)*c/100)/a,5)  # Ответ
    while (round(K * 100 - int(K * 100), 5) != 0 or K<=0 or K>=1 or a==b or a==c or b==c or round(K,2)==0):  # Пока после запятой более 1 цифры
        a, b, c = TRU(a, b, c)  # создаём новый набор
        K = (b-(100-a)*c/100)/a  # Ответ выражения меняем на новое
    txt = 'В городе ' + str(round(a,2)) + '% взрослого населения— женщины. Пенсионеры составляют ' + latex(b) + '% взрослого населения, причём доля пенсионеров среди мужчин равна ' + latex(c) + '%. Для социологического опроса выбрана случайным образом женщина, проживающий в этом городе. Найдите вероятность события «выбранная женщина является пенсионером».'
    txt = txt.replace('.0%', '%')
    ws.cell(row=i+1, column=4).value = txt
    ws.cell(row=i+1, column=6).value = K
    ws.cell(row=i+1, column=7).value = a
    ws.cell(row=i+1, column=8).value = b
    ws.cell(row=i+1, column=9).value = c

wb.save("test10.xlsx")