from sympy import *
import numpy as np
import random
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])


x = Symbol('x')
points = np.array([0, pi/2, pi, 3*pi/2, 2*pi])
func_trig = np.array([cos(x), sin(x)])

amount=2
i=2
while (i < 100 + 1):
    c = np.random.choice(list(range(-10, -3)) + list(range(1, 20)))
    n = np.random.choice(list(range(-20, -1)) + list(range(1, 20)))
    m = np.random.choice(list(range(-20, -1)) + list(range(1, 20)))
    y = np.random.choice(func_trig, size = 2, replace = false)
    eq = (n / (m * y[1] ** 2 + m * (-1) ** random.randint(1, 2) * y[1] + m * c))
    solution = solve(Eq(eq.diff(x), 0), x)
    k= []
    for elem in solution:
        k.append(eq.subs(x, elem))
    ans_1 = min(k)
    ans_2 = max(k)
    p=range(1, 3)
    if y[1]==sin(x):
        point_1, point_2 = -pi/2, pi/2
    else:
        point_1, point_2 = 0, pi

    while(ans_1 is nan or round(ans_1 * 10**amount - int(ans_1 * 10**amount), 5) != 0 ):
        c = np.random.choice(list(range(-10, -3)) + list(range(1, 20)))
        n = np.random.choice(list(range(-20, -1)) + list(range(1, 20)))
        m = np.random.choice(list(range(-20, -1)) + list(range(1, 20)))
        y = np.random.choice(func_trig, size=2, replace=false)
        eq = (n / (m * y[1] ** 2 + m * (-1) ** random.randint(1, 2) * y[1] + m * c))
        solution = solve(Eq(eq.diff(x), 0), x)
        k = []
        for elem in solution:
            k.append(eq.subs(x, elem))
        ans_1 = min(k)
        ans_2 = max(k)
        p = range(1, 3)
        if y[1] == sin(x):
            point_1, point_2 = -pi / 2, pi / 2
        else:
            point_1, point_2 = 0, pi

    txt_1 = "Найдите наименьшее значение функции $$"
    txt=txt_1 + latex(eq, ln_notation = true) + "$$ На отрезке: $$x = [" + latex(simplify(point_1))+',' +latex(simplify(point_2))+ "]$$"
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(N(ans_1)), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(N(ans_1)), 3):g}'.replace('.', ',')
    i+=1


    while (ans_1 is nan or round(ans_2 * 10**amount - int(ans_2 * 10**amount), 5) != 0):
        c = np.random.choice(list(range(-10, -3)) + list(range(1, 20)))
        n = np.random.choice(list(range(-20, -1)) + list(range(1, 20)))
        m = np.random.choice(list(range(-20, -1)) + list(range(1, 20)))
        y = np.random.choice(func_trig, size=2, replace=false)
        eq = (n / (m * y[1] ** 2 + m * (-1) ** random.randint(1, 2) * y[1] + m * c))
        solution = solve(Eq(eq.diff(x), 0), x)
        k = []
        for elem in solution:
            k.append(eq.subs(x, elem))
        ans_1 = min(k)
        ans_2 = max(k)
        p = range(1, 3)
        if y[1] == sin(x):
            point_1, point_2 = -pi / 2, pi / 2
        else:
            point_1, point_2 = 0, pi

    txt_2 = "Найдите наибольшее значение функции $$"
    txt=txt_2 + latex(eq, ln_notation = true) + "$$ На отрезке: $$x = [" + latex(simplify(point_1))+',' +latex(simplify(point_2))+ "]$$"
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(N(ans_2)), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(N(ans_2)), 3):g}'.replace('.', ',')
    i+=1

wb.save("test7.xlsx")
