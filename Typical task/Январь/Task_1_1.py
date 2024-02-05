from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

name_corp = ["«Бесконечная связь»", "«Свобода в общении»", "«Экономичный выбор»", "«Максимальная доступность»",
        "«Легко и выгодно»", "«Безлимитный обмен»", "«Ультра-коммуникации»", "«Прозрачная связь»",
        "«Связь без границ»", "«Идеальная связь»", "«Тариф для всех»", "«Связь вне ограничений»",
        "«Всегда на связи»", "«Связь без ограничений»", "«Безграничные возможности»", "«Связь по выгодной цене»",
        "«Связь для каждого»", "«Эконом-коммуникации»", "«Максимальная свобода»", "«Безлимит в общении»",
        "«Простота и доступность»", "«Бесконечный разговор»", "«Связь на выбор»", "«Ультра-доступность»",
        "«Идеальное общение»", "«Свободный выбор»", "«Экономичное общение»", "«Тариф по желанию»",
        "«Бесконечные разговоры»", "«Свобода общения»", "«Экспресс-связь»"
        ]
name_client=["Лизы", "Маши", "Даши", "Инны", "Алисы", "Дианы", "Тани", "Миланы", "Анны", "Софии"]


wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2"])
number = 100
i=2
amount=1
len_corp=len(name_corp)
len_client=len(name_client)
a = Symbol('a')
b = Symbol('b')
# c = Symbol('c')
while (i < number + 2):
    a = random.randint(10, 30)
    b = random.randint(200, 5000)
    my_name_corp = name_corp[random.randint(0, len_corp - 1)]
    my_client = name_client[random.randint(0, len_client - 1)]
    Otvet=b//a
    txt = "По тарифному плану " + my_name_corp + " компания сотовой связи каждый вечер снимает со счёта абонента " \
          + str(a) + " руб. Если на счету осталось меньше " + str(a) + " руб., то на следующее утро номер блокируют до пополнения счёта. " \
          + "Сегодня утром у "+ my_name_corp + " на счету было " + str(b) + " руб. Сколько дней (включая сегодняшний) она сможет пользоваться телефоном, не пополняя счёт?"


    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{int(b)}'
    i+=1

wb.save("test_1_1.xlsx")