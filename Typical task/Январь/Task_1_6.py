from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

name_product = ["клубники", "малины", "черники", "ежевика", "земляника",
                "винограда", "граната", "киви", "ежевика", "земляника",
                "клубники", "малины", "черники", "манго", "бананов",
                "фиников", "апельсинов", "мандаринов", "авокадо", "груш"
                ]
name_woman = ["Мама", "Маша", "Наташа", "Лера", "Алиса"]
name_man = ["Папа", "Вася", "Руслан", "Ваня", "Игорь"]
name_sex = ["купила", "купил", "она", "он"]

bills=[100, 200, 500, 1000, 2000, 5000]
wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2"])
number = 100
i=2
amount=0
a = Symbol('a')
b = Symbol('b')
len_product=len(name_product)
len_woman=len(name_woman)
len_man=len(name_man)
while (i < number + 2):
    a = random.randint(12, 20)*5
    b = random.randint(1, 4)
    c = random.randint(1, 9)*100
    n = random.randint(5, 100)*100
    Otvet = int(n-a*(b+c/1000))
    while Otvet<0 or (Otvet>100 and Otvet not in bills):
        a = random.randint(12, 20) * 5
        b = random.randint(1, 5)
        c = random.randint(1, 9) * 100
        n = random.randint(5, 100) * 100
        Otvet = int(n - a * (b + c / 1000))
    if (i%(len_man+len_woman)<len_woman):
        NAME=name_woman[i%len_woman]
    else:
        NAME = name_man[i % len_man]
    txt = "Летом килограмм "+name_product[i%len_product] + " стоит "\
           +str(a)+" рублей. "+NAME+" "+name_sex[0 if i%(len_man+len_woman)<len_woman else 1]\
           +" "+str(b)+" кг "+str(c)+" г "+name_product[i%len_product]+". Сколько рублей сдачи "\
           +name_sex[2 if i%(len_man+len_woman)<len_woman else 3]+" получит с "+str(n)+" рублей?"


    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{Otvet}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{b}'
    ws.cell(row=i, column=8).value = f'{c}'
    ws.cell(row=i, column=9).value = f'{n}'
    i+=1

wb.save("test_1_6.xlsx")