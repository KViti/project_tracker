from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

name_holiday = ["с новогодними поздравлениями",
                "с поздравлениями 8 марта",
                "с поздравлениями 23 февраля",
                "с поздравлениями 1 мая",
                "с поздравлениями 9 мая",
                "с поздравлениями 1 сентября"
                ]
name_people = ["друзьям", "подругам", "друзьям",
               "друзьям", "друзьям", "друзьям"
               ]
name_woman = ["Маша", "Наташа", "Лера", "Алиса", "Лена", "Юля", "Даша", "Катя", "Надя"]
name_woman_2 = ["Маши", "Наташи", "Леры", "Алисы", "Лены", "Юли", "Даши", "Кати", "Нади"]
name_man = ["Вася", "Руслан", "Ваня", "Игорь", "Андрей", "Даня", "Егор", "Вадим"]
name_man_2 = ["Васи", "Руслана", "Вани", "Игоря", "Андрея", "Дани", "Егора", "Вадима"]
name_sex = ["отправила", "отправил"]
RUB = ["рубль", "рубля"]
bills=[100, 200, 500, 1000, 2000, 5000]
wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2"])
number = 100
i=2
amount=0
a = Symbol('a')
b = Symbol('b')
len_holiday=len(name_holiday)
len_woman=len(name_woman)
len_man=len(name_man)
while (i < number + 2):
    a = random.randint(10, 30)
    b = random.randint(1, 2)
    c = random.randint(1, 9)*10
    n = random.randint(30, 100)
    Otvet = n-a*(b+c/100)
    while Otvet<0:
        a = random.randint(10, 30)
        b = random.randint(1, 2)
        c = random.randint(1, 9) * 10
        n = random.randint(30, 100)
        Otvet = n - a * (b + c / 100)
    Otvet = round(float(Otvet), 1)
    if Otvet == int(Otvet):
        Otvet = int(Otvet)
    if (i%(len_man+len_woman)<len_woman):
        NAME=name_woman[i%len_woman]
        NAME_2=name_woman_2[i%len_woman]
    else:
        NAME = name_man[i % len_man]
        NAME_2 =name_man_2[i % len_man]
    txt = NAME+" "+name_sex[0 if i%(len_man+len_woman)<len_woman else 1]+" SMS-сообщения "+name_holiday[i%len_holiday]+" своим "+str(a)+" "+name_people[i%len_holiday]\
          +". Стоимость одного SMS-сообщения "+str(b)+" "+RUB[0 if b==1 else 1]+" "+str(c)+" копеек. Перед отправкой сообщения на счету у "+NAME_2\
          +" было "+str(n)+" рублей. Сколько рублей останется у "+NAME_2+" после отправки всех сообщений?"

    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{Otvet}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{b}'
    ws.cell(row=i, column=8).value = f'{c}'
    ws.cell(row=i, column=9).value = f'{n}'
    i+=1

wb.save("test_1_7.xlsx")