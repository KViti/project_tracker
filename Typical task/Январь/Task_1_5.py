from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook
name_group=["В летнем легере ", "В экскурсионной группе "]
name_a = [" детей", " туристов"]
name_b = [" воспитателей. ", " сопровождающих. "]
name_c = [" из лагеря в город?", " c экскурсии в город?"]
wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2"])
number = 100
i=2
amount=0
a = Symbol('a')
b = Symbol('b')
# c = Symbol('c')
while (i < number + 2):
    a = random.randint(200, 500)
    b = random.randint(10, 40)
    c = random.randint(200, 500)
    Otvet = math.ceil((a+b)/c)
    while (a+b)/c==int((a+b)/c):
        a = random.randint(200, 500)
        b = random.randint(10, 40)
        c = random.randint(200, 500)
        Otvet = math.ceil((a + b) / c)

    txt = name_group[i%2]+str(a)+name_a[i%2]+" и "\
          +str(b)+name_b[i%2]+"В автобус помещается не более "\
          + str(c)+" пассажиров. Сколько автобусов требуется, чтобы перевезти всех "+name_c[i%2]


    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{Otvet}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{b}'
    ws.cell(row=i, column=8).value = f'{c}'
    i+=1

wb.save("test_1_5.xlsx")