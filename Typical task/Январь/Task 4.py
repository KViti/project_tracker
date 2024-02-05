from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно с точкой", "Параметр 1", "Параметр 2", "Параметр 3", "Параметр 4"])
number = 100
i=2
amount=1
a = Symbol('a')
b = Symbol('b')
c = Symbol('c')
while (i < number + 2):
    a = random.randint(1, 100)
    b = random.randint(1, 99)
    c = random.randint(1, 100)
    d = random.randint(1, 99)
    Otvet = (a*b+c*d)/(a+c)
    if(round(Otvet * 10 ** amount - int(Otvet * 10 ** amount), 5) != 0 or b==d or a==c):
        continue
    txt = 'Смешали ' + str(a)
    if a%10==1 and (a%100<10 or a%100>20):
        txt += ' литр ' + str(b) + '−процентного водного раствора некоторого вещества с '
    elif a%10<=4 and a%10!=0 and (a%100<10 or a%100>20):
        txt += ' литра ' + str(b) + '−процентного водного раствора некоторого вещества с '
    else:
        txt += ' литров ' + str(b) + '−процентного водного раствора некоторого вещества с '

    if c%10==1 and (c<10 or c>20):
        txt += str(c) + ' литром '
    else:
        txt += str(c) + ' литрами '


    txt += str(d) + '−процентного водного раствора этого же вещества. Сколько процентов составляет концентрация получившегося раствора?'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{round(float(Otvet), 3):g}'.replace(',', '.')
    ws.cell(row=i, column=7).value = f'{a}'
    ws.cell(row=i, column=8).value = f'{b}'
    ws.cell(row=i, column=9).value = f'{c}'
    ws.cell(row=i, column=10).value = f'{d}'
    i+=1

wb.save("test4.xlsx")