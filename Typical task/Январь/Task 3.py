from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2", "Параметр 3"])
number = 100
i=2
amount=0
a = Symbol('a')
b = Symbol('b')
c = Symbol('c')
while (i < number + 2):
    n = random.randint(1, 40)
    b = random.randint(1, 20)
    c = random.randint(1, 48)
    a = (c * n ** 2 - c * b * n) / b
    Otvet = n
    if (round(a * 10 ** amount - int(a * 10 ** amount), 5) != 0 or a <= 0):
        continue
    print(Otvet)
    txt = 'Заказ на ' + str(int(a))
    if a % 10 == 1 and (a % 100 < 10 or a % 100 > 20):
        txt += ' деталь первый рабочий выполняет на ' + str(c)
    elif a % 10 <= 4 and a % 10 != 0 and (a % 100 < 10 or a % 100 > 20):
        txt += ' детали первый рабочий выполняет на ' + str(c)
    else:
        txt += ' деталей первый рабочий выполняет на ' + str(c)

    if c % 10 == 1 and (c % 100 < 10 or c % 100 > 20):
        txt += ' час быстрее, чем второй. Сколько деталей в час делает первый рабочий, если известно, что он за час делает на ' \
               + str(b)
    elif c % 10 <= 4 and c % 10 != 0 and (c % 100 < 10 or c % 100 > 20):
        txt += ' часа быстрее, чем второй. Сколько деталей в час делает первый рабочий, если известно, что он за час делает на ' \
               + str(b)
    else:
        txt += ' часов быстрее, чем второй. Сколько деталей в час делает первый рабочий, если известно, что он за час делает на ' \
               + str(b)

    if b % 10 == 1 and (b % 100 < 10 or b % 100 > 20):
        txt += ' деталь больше?'
    elif b % 10 <= 4 and b % 10 != 0 and (b % 100 < 10 or b % 100 > 20):
        txt += ' детали больше?'
    else:
        txt += ' деталей больше?'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{Otvet}'
    ws.cell(row=i, column=6).value = f'{int(a)}'
    ws.cell(row=i, column=7).value = f'{b}'
    ws.cell(row=i, column=8).value = f'{c}'
    i+=1

wb.save("test3.xlsx")