from sympy import *
import random
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])

def znak(t):  # Функция для определения знака
    if t > 0:
        return '+'
    else:
        return '-'

amount = 2
number = 100
x = Symbol('x', real=True)
n = Symbol('n', real=True)
i = 2
while (i < number + 1):
    a, b, c, k = random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50)
    y = a * x ** 2 + b * x + c
    n = random.randint(2, 10)
    eq = log(y, n)+k
    solution = -b / (2 * a) # Ответ
    K = c-b*(b/(2*a))+a*(b/(2*a))**2

    if (K<0 or round(solution * 10 ** amount - int(float(solution) * 10 ** amount), 5) != 0):
        continue
    txt = "Найдите точку минимума функции $$y=\log_{" + latex(n) + '}(' + latex(y) + ')'+znak(k)+str(abs(k))+'.$$'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(solution), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(solution), 3):g}'.replace('.', ',')
    i+=2

i = 3
while (i < number + 2):
    a, b, c, k = -random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50)
    y = a * x ** 2 + b * x + c
    n=random.randint(2, 10)
    eq = log(y, n)
    solution = -b / (2 * a)# Ответ
    K = c-b*(b/(2*a))+a*(b/(2*a))**2
    if (K<0 or round(solution * 10 ** amount - int(float(solution) * 10 ** amount), 5) != 0):
        continue

    max_point_1 = int(solution) - random.randint(3, 10)
    max_point_2 = int(solution) + random.randint(3, 10)
    txt = "Найдите точку максимума функции $$y=\log_{" + latex(n) + '}(' + latex(y) + ')'+znak(k)+str(abs(k))+'.$$'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(solution), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(solution), 3):g}'.replace('.', ',')
    i+=2

wb.save("test3.xlsx")
