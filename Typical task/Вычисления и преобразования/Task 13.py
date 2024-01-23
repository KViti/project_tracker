from sympy import *
import numpy as np
import random
from openpyxl import Workbook
import math

def reduce_fraction_Denominator(n, m):  # Функция сокращающая дробь выводит знаменатель
    k = math.gcd(n, m)
    return (m // k)


wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])


def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return float(str(t*10**amount).split('.')[0])-int(str(t*10**amount).split('.')[0])==0

amount = 2
number = 100
i = 2
while (i < number + 2):

    osnova_1 = [(2 * i + 1) for i in range(1, 5)]
    osnova_2 = [(2 ** i) for i in range(1, 3)]
    p = np.random.choice(osnova_1)
    k = np.random.choice(osnova_2)
    otvet = p/k  # Ответ

    n = random.randint(2, 10)
    a = n ** k
    b = n ** p

    # a == b or b == 1 or b > 200 or a > 200 or b < 0.01 or a < 0.01 or
    while (abs(a)>2000 or abs(b)>2000 or a==zoo or a==nan or b==zoo or b==nan or Prov(otvet, amount)):

        p = np.random.choice(osnova_1)
        k = np.random.choice(osnova_2)
        otvet = p / k  # Ответ


        n = random.randint(2, 5)
        a = n ** k
        b = n ** p

    txt = "Найдите значение выражения $$\log_{" + str(a) + '}' + str(b)
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    print(i)
    i += 2

i = 3
while (i < number + 2):
    p = random.randint(1, 20)
    k = random.randint(2, 20)
    otvet = p/k  # Ответ

    n = random.randint(2, 5)
    a = n ** k
    b = n ** p
    # a == b or b == 1 or b > 200 or a > 200 or b < 0.01 or a < 0.01 or
    while (abs(a)>2000 or abs(b)>2000 or a==zoo or a==nan or b==zoo or b==nan or Prov(otvet, amount) or reduce_fraction_Denominator(p, k) in [3, 6, 7, 9]):
        p = random.randint(1, 20)
        k = random.randint(2, 20)
        otvet = p / k  # Ответ
        n = random.randint(2, 10)
        a = n ** k
        b = n ** p


    txt = "Найдите значение выражения $$\log_{" + latex(a) + '}' + latex(b)
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    print(i)
    i += 2
wb.save("test13.xlsx")
