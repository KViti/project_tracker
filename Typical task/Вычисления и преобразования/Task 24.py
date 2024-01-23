from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

amount = 100
wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])
osnov=[2, 3, 4, 5, 6, 7]
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]
osnova_3 = [2, 4, 5, 10, 20, 25, 50, 100]
chet = [2*i for i in range(1, 10)]
i=2
while (i< amount + 2):
    a = random.randint(2, 10)
    b = random.randint(2, 10)
    c = random.randint(2, 10)
    k = random.randint(2, 10)

    chetn = np.random.choice(chet)
    m = k ** chetn
    otvet = b + c ** (chetn / 2)
    while (abs(m) > 1000 or abs(otvet) > 1000):
        a = random.randint(2, 10)
        b = random.randint(2, 10)
        c = random.randint(2, 10)
        k = random.randint(2, 10)
        chetn = np.random.choice(chet)
        m=k**chetn
        otvet = b+c**(chetn/2)

    txt = "Найдите значение выражения $$" +latex(a)+'^{\log_{' + latex(a) + '}' + latex(b) +'}+'+ latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    txt = "Найдите значение выражения $$" + latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}+'+latex(a)+ '^{\log_{' + latex(a) + '}' + latex(b) +'}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1


    otvet = b - c ** (chetn / 2)
    txt = "Найдите значение выражения $$" +latex(a)+'^{\log_{' + latex(a) + '}' + latex(b) +'}-'+ latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    otvet = c ** (chetn / 2)-b
    txt = "Найдите значение выражения $$" + latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}-'+latex(a)+ '^{\log_{' + latex(a) + '}' + latex(b) +'}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1











    number_1 = random.randint(2, 10)
    number_2 = random.randint(2, 10)
    otvet = number_1 * c ** (chetn / 2) + b
    while (abs(m) > 1000 or abs(otvet) > 1000):
        number_1 = random.randint(2, 10)
        a = random.randint(2, 10)
        b = random.randint(2, 10)
        c = random.randint(2, 10)
        k = random.randint(2, 10)
        chetn = np.random.choice(chet)
        m=k**chetn
        otvet = number_1 * c ** (chetn / 2) + b

    txt = "Найдите значение выражения $$" + latex(number_1) +'\cdot'+ latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}+'+latex(a)+ '^{\log_{' + latex(a) + '}' + latex(b) +'}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1


    otvet = c ** ((chetn / 2)*number_1) + b
    while (abs(m) > 1000 or abs(otvet) > 1000):
        number_1 = random.randint(2, 10)
        a = random.randint(2, 10)
        b = random.randint(2, 10)
        c = random.randint(2, 10)
        k = random.randint(2, 10)
        chetn = np.random.choice(chet)
        m=k**chetn
        otvet = c ** ((chetn / 2)*number_1) + b
    txt = "Найдите значение выражения $$" + latex(int(m)) + '^{\log_{' + latex(k) + '}(' + latex(sqrt(c)) + ')^{'+latex(number_1) + '}}+'+latex(a)+ '^{\log_{' + latex(a) + '}' + latex(b) +'}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1


    otvet = c ** (chetn / 2) - number_1 * b
    while (abs(m) > 1000 or abs(otvet) > 1000):
        number_1 = random.randint(2, 10)
        a = random.randint(2, 10)
        b = random.randint(2, 10)
        c = random.randint(2, 10)
        k = random.randint(2, 10)
        chetn = np.random.choice(chet)
        m=k**chetn
        otvet = c ** (chetn / 2) - number_1 * b

    txt = "Найдите значение выражения $$" + latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}-'+latex(number_1) +'\cdot'+latex(a)+ '^{\log_{' + latex(a) + '}' + latex(b) +'}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1


    otvet = c ** (chetn / 2) - b**number_1
    while (abs(m) > 1000 or abs(otvet) > 1000):
        number_1 = random.randint(2, 10)
        a = random.randint(2, 10)
        b = random.randint(2, 10)
        c = random.randint(2, 10)
        k = random.randint(2, 10)
        chetn = np.random.choice(chet)
        m=k**chetn
        otvet = c ** (chetn / 2) - b**number_1
    txt = "Найдите значение выражения $$" + latex(int(m)) + '^{\log_{' + latex(k) + '}' + latex(sqrt(c)) + '}-' +latex(a)+ '^{\log_{' + latex(a) + '}' + latex(b) + '^{' +latex(number_1) +'}}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1



wb.save("test24.xlsx")