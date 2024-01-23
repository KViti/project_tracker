from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

amount = 100
wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])
osnov=[2, 3, 4, 5, 6, 7]
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]
osnova_3 = [2, 4, 5, 10, 20, 25, 50, 100]
i=2
while (i< amount + 2):
    n = random.randint(2, 10)  # основание
    k = random.randint(2, 10)
    s = random.randint(2, 10)
    p = random.randint(2, 10)
    b = np.random.choice(osnova_2) / 10
    t = np.random.choice(osnova_3)
    a_b = k ** s
    otvet = s / t
    a = a_b * b
    while (abs(a_b) > 1000 or abs(t**p) > 1000 or abs(n**p) > 1000 or Prov(otvet, amount) or t**p==n**p):
        n = random.randint(2, 10)  # основание
        k = random.randint(2, 10)
        s = random.randint(2, 10)
        p = random.randint(2, 10)
        b = np.random.choice(osnova_2) / 10
        t = np.random.choice(osnova_3)
        a_b = k ** s
        otvet = s / t
        a = a_b * b

    txt = "Найдите значение выражения $$\\frac{\log_{" + latex(k) + '}' + latex(a) +'-\log_{'+ latex(k) + '}' + latex(b) + '}{'+ latex(n) +'^{\log_{' + latex(n**p) + '}' + latex(int(t**p)) + '}}'
    txt = txt.replace('.0', '')
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    n = random.randint(2, 10)  # основание
    k = random.randint(2, 10)
    s = random.randint(2, 10)
    p = random.randint(2, 10)
    b = np.random.choice(osnova_2) / 10
    t = np.random.choice(osnova_3)
    a_b = k**s
    otvet = s / t
    a = a_b/b
    while (abs(a_b) > 1000 or abs(t**p) > 1000 or abs(n**p) > 1000 or Prov(otvet, amount) or t**p==n**p):
        n = random.randint(2, 10)  # основание
        k = random.randint(2, 10)
        s = random.randint(2, 10)
        p = random.randint(2, 10)
        b = np.random.choice(osnova_2) / 10
        t = np.random.choice(osnova_3)
        a_b = k ** s
        otvet = s / t
        a = a_b / b

    txt = "Найдите значение выражения $$\\frac{\log_{" + latex(k) + '}' + latex(a) +'+\log_{'+ latex(k) + '}' + latex(b) + '}{'+ latex(n) +'^{\log_{' + latex(n**p) + '}' + latex(int(t**p)) + '}}'
    txt = txt.replace('.0', '')
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

wb.save("test23.xlsx")