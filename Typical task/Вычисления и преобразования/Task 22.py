from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])
osnov=[2, 3, 4, 5, 6, 7]
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]
osnova_3 = [2, 4, 5, 10, 20, 25, 50, 100]
i=2
while (i< amount + 2):
    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 10)
    otvet=n
    txt = "Найдите значение выражения $${\left(" + latex(n) +"^{\log_{" + latex(a) + '}{' + latex(b) + '}}\\right)^{\log_{' + latex(b) + '}' + latex(a) + '}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 10)
    otvet = n ** k
    while (abs(otvet) > 1000 or a==b):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)
        b = random.randint(2, 20)
        k = random.randint(2, 10)
        otvet = n ** k
    txt = "Найдите значение выражения $${\left(" + latex(n) +"^{\log_{" + latex(a) + '}{' + latex(b) + '}}\\right)^{'+latex(k)+'\log_{' + latex(b) + '}' + latex(a) + '}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 10)
    otvet = n ** k
    while (abs(otvet) > 1000 or a==b):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)
        b = random.randint(2, 20)
        k = random.randint(2, 10)
        otvet = n ** k
    txt = "Найдите значение выражения $${\left(" + latex(n) +"^{"+latex(k)+"{\log_{" + latex(a) + '}{' + latex(b) + '}}}\\right)^{\log_{' + latex(b) + '}{' + latex(a) + '}}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 10)
    otvet = n ** (k * k)
    while (abs(otvet) > 1000 or a==b):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)
        b = random.randint(2, 20)
        k = random.randint(2, 10)
        otvet = n ** (k * k)
    txt = "Найдите значение выражения $${\left(" + latex(n) +"^{\log_{" + latex(a) + '}{' + latex(b) +"}^{"+latex(k)+ '}}\\right)^{'+latex(k)+'\log_{' + latex(b) + '}{' + latex(a) + '}}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 10)
    otvet = n ** k
    t = random.randint(2, 10)
    while (abs(otvet) > 1000 or a==b):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)
        b = random.randint(2, 20)
        k = random.randint(2, 10)
        otvet = n ** k
    txt = "Найдите значение выражения $${\left(" + latex(n) +"^{\log_{" + latex(a) + '}' + "^{" +latex(t)+'}{' + latex(b) + '}}\\right)^{'+latex(k)+'\log_{' + latex(b) + '}' +"^{"+latex(t)+'}{'+latex(a) + '}}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 10)
    otvet = n ** k
    while (abs(otvet) > 1000 or a==b):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)
        b = random.randint(2, 20)
        k = random.randint(2, 10)
        otvet = n ** k
    txt = "Найдите значение выражения $${\left(\left(" + latex(n) +"^{\log_{" + latex(a) + '}{' + latex(b) + '}}\\right)^{\log_{' + latex(b) + '}{' + latex(a) + '}}\\right)^{' + latex(k) +'}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

    a = random.randint(2, 20)
    b = random.randint(2, 20)
    k = random.randint(2, 3)
    otvet = random.randint(2, 20)
    n = otvet ** k
    while (abs(n) > 1000 or a==b):
        a = random.randint(2, 20)
        b = random.randint(2, 20)
        k = random.randint(2, 3)
        otvet = random.randint(2, 20)
        n = otvet**k
    txt = "Найдите значение выражения $$\left(" + latex(n) +"^{\log_{" + latex(a)+ '^{' + latex(k) + '}}{' + latex(b) + '}}\\right)^{\log_{' + latex(b) + '}{' + latex(a) + '}}'
    txt = txt.replace('.0', '') + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1

wb.save("test22.xlsx")