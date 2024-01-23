from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return float(str(t*10**amount).split('.')[0])-int(str(t*10**amount).split('.')[0])==0

amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])

i=2
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]
osnova_3 = [2, 4, 5, 10, 20, 25, 50, 100]

while (i< amount + 2):
    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)  # Число
    b = random.randint(2, 20)  #
    m = n**a*b
    k = random.randint(2, 20)
    otvet = k  #

    while (abs(m)>200):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)  # Число
        b = random.randint(2, 20)  #
        m = n ** a * b
    txt = "Найдите значение выражения $$\\frac{" + latex(k) + "\log_{" + latex(n) + '}' + latex(m) + '}{' + latex(a) + '+\log_{' + latex(n) + '}' + latex(int(b)) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1
    txt = "Найдите значение выражения $$\\frac{" + latex(k*a) + '+' + latex(k) + '\log_{' + latex(n) + '}' + latex(int(b)) + '}{\log_{' + latex(n) + '}' + latex(m) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1

    n = random.randint(2, 20)  # основание
    a = random.randint(2, 20)  # Число
    b = random.randint(2, 20)  #
    m = n**a*b
    k = np.random.choice(osnova_3)
    otvet = 1/k  #

    while (abs(m)>200):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 20)  # Число
        b = random.randint(2, 20)  #
        m = n ** a * b
    txt = "Найдите значение выражения $$\\frac{\log_{" + latex(n) + '}' + latex(m) + '}{' + latex(int(k*a)) + '+' + latex(int(k)) + '\log_{' + latex(n) + '}' + latex(int(b)) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1
    txt = "Найдите значение выражения $$\\frac{" + latex(a) + '+\log_{' + latex(n) + '}' + latex(b) + '}{' + latex(int(k)) + '\log_{' + latex(n) + '}' + latex(m) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1



    n = random.randint(2, 20)  # основание
    a = random.randint(2, 100)  # Число
    b = np.random.choice(osnova_3)  #
    otvet = a/b  #

    while (Prov(otvet, amount)):
        n = random.randint(2, 20)  # основание
        a = random.randint(2, 100)  # Число
        b = np.random.choice(osnova_3)  #
        otvet = a / b
    txt = "Найдите значение выражения $$\\frac{" + latex(a) + '}{' + latex(n) +'^{\log_{' + latex(n) + '}' + latex(int(b)) + '}}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1

    n = random.randint(2, 20)  # основание
    b = random.randint(2, 100)  # Число
    a = np.random.choice(osnova_3)  #
    otvet = b/a  #

    while (Prov(otvet, amount)):
        n = random.randint(2, 20)  # основание
        b = random.randint(2, 100)  # Число
        a = np.random.choice(osnova_3)  #
        otvet = b/a

    txt = "Найдите значение выражения $$\\frac{" + latex(n) +'^{\log_{' + latex(n) + '}' + latex(int(b)) + '}}{' + latex(int(a)) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1

wb.save("test18.xlsx")