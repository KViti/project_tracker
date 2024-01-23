from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0


amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])
osnov=[2, 3, 4, 5, 6, 7]
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]
osnova_3 = [2, 4, 5, 10, 20, 25, 50, 100]
i=2
while (i< amount + 2):
    n = np.random.choice(osnov)  # основание
    p = np.random.choice(osnov)  # основание
    a = np.random.choice(osnova_2)
    b = random.randint(1, 3)
    c = random.randint(1, 3)
    otvet = (c-b)/a
    number1 = n**a
    number2 = p**(n**b)
    number3 = p**(n**c)
    while (abs(number3)>1000 or Prov(otvet, amount) or abs(number2)>1000 or number2==0 or number3==0):
        n = np.random.choice(osnov)  # основание
        p = np.random.choice(osnov)  # основание
        a = np.random.choice(osnova_2)
        b = random.randint(1, 3)
        c = random.randint(1, 3)
        otvet = (c - b) / a
        number1 = n ** a
        number2 = p ** (n ** b)
        number3 = p ** (n ** c)

    txt = "Найдите значение выражения $$\log_{" + latex(float(number1)) + '}\log_{' + latex(float(number2)) + '}' + latex(float(number3))
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1

    otvet = (b-c) / a
    while (abs(number3)>1000 or Prov(otvet, amount) or abs(number2)>1000 or number2==0 or number3==0):
        n = np.random.choice(osnov)  # основание
        p = np.random.choice(osnov)  # основание
        a = np.random.choice(osnova_2)
        b = random.randint(1, 3)
        c = random.randint(1, 3)
        otvet = (b-c) / a
        number1 = n ** a
        number2 = p ** (n ** b)
        number3 = p ** (n ** c)

    txt = "Найдите значение выражения $$\log_{\\frac{1}{" + latex(float(number1)) + '}}\log_{' + latex(float(number2)) + '}' + latex(float(number3))
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1

    n = np.random.choice(osnov)  # основание
    p = np.random.choice(osnov)  # основание
    a = np.random.choice(osnova_1)
    c = random.randint(1, 3)
    otvet = (c-1) / a
    number1 = n ** a
    number2 = p ** n
    number3 = p ** (n ** c)
    while (abs(number3) > 1000 or Prov(otvet, amount) or abs(number2) > 1000 or number2 == 0 or number3 == 0):
        n = np.random.choice(osnov)  # основание
        p = np.random.choice(osnov)  # основание
        a = np.random.choice(osnova_1)
        c = random.randint(1, 3)
        otvet = (c - 1) / a
        number1 = n ** a
        number2 = p ** n
        number3 = p ** (n ** c)

    txt = "Найдите значение выражения $$\log_{" + latex(float(number1)) + '}\log_{' + latex(float(number2)) + '}' + latex(float(number3))
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1



    # n = np.random.choice(osnov)  # основание
    # p = np.random.choice(osnov)  # основание
    # c = random.randint(1, 5)
    # otvet = c-1
    # number1 = n
    # number2 = p ** n
    # number3 = p ** (n ** c)
    # while (abs(number3) > 1000 or Prov(otvet, amount) or abs(number2) > 1000 or number2 == 0 or number3 == 0):
    #     n = np.random.choice(osnov)  # основание
    #     p = np.random.choice(osnov)  # основание
    #     c = random.randint(1, 3)
    #     otvet = (c - 1)
    #     number1 = n
    #     number2 = p ** n
    #     number3 = p ** (n ** c)
    #
    # txt = "Найдите значение выражения $$\log_{" + latex(float(number1)) + '}$$\log_{' + latex(float(number2)) + '}' + latex(float(number3)) + '}'
    # txt = txt.replace('.0', '')
    # txt = txt.replace("\\left(", "")
    # txt = txt.replace("\\right)", "") + '.$$'
    # ws.cell(row=i, column=4).value = txt
    # ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    # ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    # i += 1

wb.save("test20.xlsx")