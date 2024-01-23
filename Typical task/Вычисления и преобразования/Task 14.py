from sympy import *
import numpy as np
import random
from openpyxl import Workbook
import math

def reduce_fraction_Denominator(n, m):  # Функция сокращающая дробь выводит знаменатель
    k = math.gcd(n, m)
    return (m // k)


wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])


def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return float(str(t*10**amount).split('.')[0])-int(str(t*10**amount).split('.')[0])==0

amount = 2
number = 100
i = 2
while (i < number + 2):

    n = random.randint(2, 10)
    k = random.randint(2, 10)

    a = n ** k
    b = random.randint(2, 10)
    otvet=b**k # Ответ

    # a == b or b == 1 or b > 200 or a > 200 or b < 0.01 or a < 0.01 or
    while (abs(a)>200 or otvet>1000 or n==b):
        k = random.randint(2, 10)
        a = n ** k
        b = random.randint(2, 10)
        otvet = b ** k  # Ответ

    txt = "Найдите значение выражения $${" + str(a) +'^{\log_{' + str(n) + '}' + str(b) +'}'
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 4
i = 3
while (i < number + 2):

    n = random.randint(2, 10)
    k = random.randint(2, 10)
    c = random.randint(2, 10)
    a = n ** k
    b = random.randint(2, 10)
    otvet = c * b ** k  # Ответ

    # a == b or b == 1 or b > 200 or a > 200 or b < 0.01 or a < 0.01 or
    while (abs(a)>200 or otvet>1000 or n==b):
        k = random.randint(2, 10)
        a = n ** k
        b = random.randint(2, 10)
        otvet = c * b ** k  # Ответ

    txt = "Найдите значение выражения $${" + str(c) + latex('*') +str(a) +'^{\log_{' + str(n) + '}' + str(b) +'}'
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 4

i = 4
while (i < number + 2):
    osnova_1 = [(2 * i) for i in range(1, 10)]
    p = np.random.choice(osnova_1)
    n = random.randint(2, 10)
    k = random.randint(2, 10)

    a = n ** p
    otvet=k**(p/2) # Ответ

    # a == b or b == 1 or b > 200 or a > 200 or b < 0.01 or a < 0.01 or
    while (a==0 or abs(a)>200 or otvet>1000 or (k in [i**2 for i in range(1, 10)])):
        p = np.random.choice(osnova_1)
        n = random.randint(2, 10)
        k = random.randint(2, 10)
        a = n ** p
        otvet=k**(p/2) # Ответ

    txt = "Найдите значение выражения $${" + str(a) +'^{\log_{' + str(n) + '}' + latex(sqrt(k)) +'}'
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 4
i = 5
while (i < number + 2):

    n = random.randint(2, 10)
    k = random.randint(2, 10)
    c = random.randint(2, 3)
    a = n ** k
    b = random.randint(2, 10)
    otvet = b ** (k * c)  # Ответ

    # a == b or b == 1 or b > 200 or a > 200 or b < 0.01 or a < 0.01 or
    while (abs(a)>200 or otvet>1000 or n==b):
        k = random.randint(2, 3)
        a = n ** k
        b = random.randint(2, 10)
        otvet = b ** (k*c)  # Ответ

    txt = "Найдите значение выражения $${" + str(a) +'^{'+str(c)+'\log_{' + str(n) + '}' + str(b) +'}'
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 4
wb.save("test14.xlsx")
