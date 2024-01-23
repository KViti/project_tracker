from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def znak(t)->str:  # Функция для определения знака
    if t > 0:
        return '+'
    else:
        return '-'

def znak_0(t)->int:  # Функция для определения знака
    if t > 0:
        return 1
    else:
        return -1

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) != 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return int(str(t*10**amount).split('.')[0])==0

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])

amount=2
number=100
x = Symbol('x', real=True)
a = Symbol('a', real=True)
b = Symbol('b', real=True)
m = Symbol('k', real=True)
n = Symbol('n', real=True)
a_b= Symbol('a_b', real=True)
otvet= Symbol('otvet', real=True)
points=np.array([1, -1, 2])
osnova=np.array([1/2, 1/5])
i=2
while (i < number + 2):
    # h=np.random.choice(osnova)
    # print(h)
    m = random.randint(2, 20)
    n = random.randint(2, 20)
    otvet = (-1) ** random.randint(0, 1)*random.randint(1, 100)/100 + random.randint(0, 1)*random.randint(1, 10)
    t=random.randint(2, 5)
    b = n ** t
    a = m ** (otvet / t)
    while n==m or m==b or n==a or a>200:
        m = random.randint(2, 20) # основание 1
        n = random.randint(2, 20) # основание 2
        b = n ** t
        a = m**(otvet/t)
    if (Prov(a, amount) or Prov(round(b, 5), amount) or Prov(m, amount) or Prov(n, amount) or a>200 or b>200 or a<=0):
        continue
    txt = "Найдите значение выражения $$\log_{" + latex(m) + '}{' + latex(b) + '}\cdot\log_{' + latex(n) + '}' + latex(a)
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=2
wb.save("test12.xlsx")

i=3
while (i < number + 2):
    n = random.randint(2, 20)  # основание 1
    k = random.randint(2, 20)
    otvet = (-1) ** random.randint(0, 1)*random.randint(1, 100)/100 + random.randint(0, 1)*random.randint(1, 10)
    p=(-1) ** random.randint(0, 1)*random.randint(2, 5)
    m = n ** (p * otvet)  # основание 2
    t = k**p

    while n==m or abs(p)>200 or n==k or m==t:
        n = random.randint(2, 20)  # основание 1
        k = random.randint(2, 20)
        otvet = (-1) ** random.randint(0, 1)*random.randint(1, 100)/100 + random.randint(0, 1)*random.randint(1, 10)
        p = (-1) ** random.randint(0, 1) * random.randint(2, 5)
        m = n ** (p * otvet)  # основание 2
        t = k ** p

    if (Prov(n, amount) or Prov(m, amount) or Prov(t, amount) or m>200 or n>200 or t>200):
        continue
    txt = "Найдите значение выражения $$\\frac{\log_{" + latex(n) + '}' + latex(k) + '}{\log_{' + latex(m) + '}' + latex(t)
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=2
wb.save("test12.xlsx")