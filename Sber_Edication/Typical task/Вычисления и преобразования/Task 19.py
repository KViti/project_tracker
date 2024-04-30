from sympy import *
import numpy as np
import random
from openpyxl import Workbook

amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])

i=2
while (i< amount + 2):
    number1 = random.randint(2, 20)  # основание
    number2 = random.randint(2, 20)  # основание
    number3 = random.randint(2, 20)  # основание
    otvet = number1**number2*number3
    while otvet>2000:
        number1 = random.randint(2, 20)  # основание
        number2 = random.randint(2, 20)  # основание
        number3 = random.randint(2, 20)  # основание
        otvet = number1 ** number2 * number3

    txt = "Найдите значение выражения $$" + latex(number1) +'^{'+latex(number2)+'+\log_{' + latex(number1) + '}' + latex(number3) + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=1

wb.save("test19.xlsx")