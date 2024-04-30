from sympy import *
import numpy as np
import random
from openpyxl import Workbook
import math

def reduce_fraction_Denominator(n, m):  # Функция сокращающая дробь выводит знаменатель
    k = math.gcd(n, m)
    return (m // k)


wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])


def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return float(str(t*10**amount).split('.')[0])-int(str(t*10**amount).split('.')[0])==0

amount = 2
number = 100
i = 2
while (i < number + 2):
    b = random.randint(2, 20)  #
    p = random.randint(2, 5) #
    n = random.randint(2, 10) # Основание логарифма
    k = random.randint(2, 10) # Основание степени
    if 5 <= k <= 10:
        p = random.randint(2, 3)  #
    a = n ** p * b
    otvet=k**p # Ответ

    txt = "Найдите значение выражения $$\\frac{" + str(k) +'^{\log_{' + str(n) + '}' + latex(a) +'}}{' + str(k) +'^{\log_{' + str(n) + '}' + latex(b) +'}'
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 2

i=3
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]

while (i < number + 2):

    n = random.randint(2, 20)  # основание 1
    a = random.randint(2, 20)  # основание 1

    k = np.random.choice(osnova_1)
    t = np.random.choice(osnova_2)

    b = a**t
    otvet = 1/(t*k)
    b = a**t  # основание 2
    while (Prov(otvet, amount) or b==n or abs(b)>200 or t*k in [8, 16, 40]):
        n = random.randint(2, 20)  # основание 1
        a = random.randint(2, 20)  # основание 1

        k = np.random.choice(osnova_1)
        t = np.random.choice(osnova_2)

        b = a ** t
        otvet = 1 / (t * k)

    txt = "Найдите значение выражения $$\\frac{\log_{" + latex(n) + '}' + '\sqrt['+latex(int(k))+']{'+ latex(a)  + '}}{\log_{' + latex(n) + '}' + latex(int(b))
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '}.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=2
wb.save("test15.xlsx")
