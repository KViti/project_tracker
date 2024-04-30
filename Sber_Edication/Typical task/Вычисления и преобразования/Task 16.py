from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return float(str(t*10**amount).split('.')[0])-int(str(t*10**amount).split('.')[0])==0

amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])
i = 2

while (i< amount + 2):
    n = random.randint(2, 10) # Основание логарифма_1
    m = random.randint(2, 10)  # Основание логарифма_2
    n_m = n*m # Множитель

    if i % 2 == 0:
        k = random.randint(2, 10)  # Множитель
    else:
        k = 1

    if i % 4 >1:
        l = random.randint(2, 10)  # Множитель
    else:
        l = 1
    if k==1:
        txt = "Найдите значение выражения $$(" + latex(k) + "-"
    else:
        txt = "Найдите значение выражения $$(" + latex(k) + "-" + latex(k)

    if n==10:
        txt = txt +"\lg" + latex(n_m) + ")("
    else:
        txt = txt +"\log_{" + latex(n) + '}' + latex(n_m) + ")("

    if l==1:
        txt = txt + latex(l) + "-"
    else:
        txt = txt + latex(l) + "-" + latex(l)

    if m==10:
        txt = txt +"\lg" + latex(n_m) + ")"
    else:
        txt = txt +"\log_{" + latex(m) + '}' + latex(n_m) + ")"

    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(k*l), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(k*l), 3):g}'.replace('.', ',')
    i += 1

wb.save("test16.xlsx")