from sympy import *
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0

def Proverka(t, amount):  # Функция для определения количества знаков после запятой
    return float(str(t*10**amount).split('.')[0])-int(str(t*10**amount).split('.')[0])==0

amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])

i=2
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]

while (i< amount + 2):
    n = random.randint(2, 20)  # основание 1
    otvet = random.randint(1, 10)
    k = random.randint(2, 20)
    p = otvet*k  #

    while (abs(p)>200):
        n = random.randint(2, 20)  # основание 1
        otvet = random.randint(1, 10)
        k = random.randint(2, 20)
        p = otvet * k  #
    if k==2:
        txt = "Найдите значение выражения $$"+ latex(p) +"\log_{" + latex(n) + '}' + '\sqrt{'+ latex(n)  + '}'
    else:
        txt = "Найдите значение выражения $$"+ latex(p) +"\log_{" + latex(n) + '}' + '\sqrt['+latex(int(k))+']{'+ latex(n)  + '}'
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=3

i=3
while (i< amount + 2):
    n = random.randint(2, 20)  # основание 1
    k = random.randint(2, 20)
    p = random.randint(2, 20)  #
    otvet = p * k

    while (abs(otvet)>200):
        n = random.randint(2, 20)  # основание 1
        k = random.randint(2, 20)
        p = random.randint(2, 20)  #
        otvet = p * k
    if k==2:
        txt = "Найдите значение выражения $$" + latex(p) + "\log_{" + '\sqrt{' + latex(n) + '}}' + latex(n)
    else:
        txt = "Найдите значение выражения $$" + latex(p) + "\log_{" + '\sqrt[' + latex(int(k)) + ']{' + latex(n) + '}}' + latex(n)
    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=3

i=4
while (i< amount + 2):
    n = random.randint(2, 20)  # основание 1
    m = random.randint(2, 10)
    k = random.randint(2, 20)
    otvet = 1 / k**m
    while (abs(otvet)>200 or k**m not in [2, 4, 5, 10, 20, 25, 50, 100]):
        n = random.randint(2, 20)  # основание 1
        m = random.randint(2, 10)
        k = random.randint(2, 20)
        otvet = 1 / k ** m
    if k==2:
        txt = "Найдите значение выражения $$\log^" + latex(m) + "_{" + latex(n) + '}' + '\sqrt{' + latex(n) + '}'
    else:
        txt = "Найдите значение выражения $$\log^" + latex(m) + "_{" + latex(n) + '}' + '\sqrt[' + latex(int(k)) + ']{' + latex(n) + '}'

    txt = txt.replace('.0', '')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=3


wb.save("test17.xlsx")