from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def znak(t)->str:  # Функция для определения знака
    if t > 0:
        return '+'
    else:
        return '-'

def znak_0(t)->int:  # Функция для определения знака
    if t > 0:
        return 1
    else:
        return -1

def Prov(t, amount):  # Функция для определения знака
    return round(t * 10 ** amount - int(float(t) * 10 ** amount), 5) != 0

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])

amount=2
number=100
x = Symbol('x', real=True)
a = Symbol('a', real=True)
b = Symbol('b', real=True)
m = Symbol('m', real=True)
n = Symbol('n', real=True)
a_b= Symbol('a_b', real=True)
otvet= Symbol('otvet', real=True)
points=np.array([1, -1, 2])
i=2
while (i < number + 2):
    k = (-1) ** random.randint(0, 1)
    h=np.random.choice(points)
    n = random.randint(2, 10)
    otvet = (-1) ** random.randint(0, 1)*random.randint(1, 100)/10
    t=(-1) ** random.randint(0, 1)*random.randint(2, 10)
    b=n**t
    # a = (n ** otvet / n ** t ** znak_0(k))
    a = (n ** (otvet - t * znak_0(k)))
    while a==0:
        t = (-1) ** random.randint(0, 1) * random.randint(2, 5)
        b = n ** t
        # a = (n ** otvet / n ** t **
        a = (n ** (otvet - t * znak_0(k)))
    a=a**int(h)
    if (a == zoo or a == nan or b == zoo or b == nan or Prov(a, amount) or Prov(b, amount) or Prov(n**int(h), amount) or a>200 or b>200 or a<0 or b<0 or round(float(b), 3)==0 or round(float(a), 3)==0 or a==b):
        continue
    txt = "Найдите значение выражения $$\log_{" + latex(simplify(n ** int(h))) + '}{' + f'{round(float(a), 3):g}' + '}' + znak(k) + '\log_{' + latex(n) + '}{' + f'{round(float(b), 3):g}'
    txt = txt.replace('.', ',') + '}.$$'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=2
wb.save("test11.xlsx")
i=3
while (i < number + 2):
    j=random.randint(2, 3)
    k = (-1) ** random.randint(0, 1)
    h=np.random.choice(points)
    n = random.randint(2, 10)
    otvet = (-1) ** random.randint(0, 1)*random.randint(1, 100)/10
    t=(-1) ** random.randint(0, 1)*random.randint(2, 10)
    b=n**t
    # a = (n ** otvet / n ** (j * t * znak_0(k)))
    a = (n ** (otvet - j * t * znak_0(k)))
    while a==0:
        t = (-1) ** random.randint(0, 1) * random.randint(2, 5)
        b = n ** t
        # a = (n ** otvet / n ** (j * t * znak_0(k)))
        a = (n ** (otvet - j * t * znak_0(k)))
    a=a**int(h)
    if (a == zoo or a == nan or b == zoo or b == nan or Prov(a, amount) or Prov(b, amount) or Prov(n**int(h), amount) or a>200 or b>200 or a<0 or b<0 or round(float(b), 3)==0 or round(float(a), 3)==0  or a==b):
        continue
    txt = "Найдите значение выражения $$\log_{" + latex(simplify(n ** int(h))) + '}{' + f'{round(float(a), 3):g}' + '}' + znak(k) + latex(j) +'\log_{' + latex(n) + '}{' + f'{round(float(b), 3):g}'
    txt = txt.replace('.', ',') + '}.$$'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i+=2
wb.save("test11.xlsx")