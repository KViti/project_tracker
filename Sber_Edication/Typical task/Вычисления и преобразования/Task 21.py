from sympy import *
import numpy as np
import random
from openpyxl import Workbook

def Prov(t, amount):  # Функция для определения количества знаков после запятой
    return round((t * 10 ** amount - int(t) * 10 ** amount), 5) == 0


amount = 100

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Правильно"])
osnov=[2, 3, 4, 5, 6, 7]
osnova_1 = [2, 4, 5, 10]
osnova_2 = [1, 2, 4, 5]
osnova_3 = [2, 4, 5, 10, 20, 25, 50, 100]
i=2
while (i< amount + 2):
    n = np.random.choice(osnov)  # основание
    b = np.random.choice(osnova_3)
    if i%5==0:
        otvet = 0
    else:
        otvet = random.randint(1, 10)
    a_b = n**otvet
    k = random.randint(2, 20)
    a = a_b
    if i%5==0:
        otvet = 0
    while (abs(a) > 1000 or a==0 or Prov(a, amount)):
        b = np.random.choice(osnova_3)
        if i % 5 == 0:
            otvet = 0
        else:
            otvet = random.randint(1, 10)
        a_b = n ** otvet
        k = random.randint(2, 20)
        a = a_b / b
    if k==10:
        txt = "Найдите значение выражения $$\\frac{\lg" + latex(a) + '}{\lg' + latex(int(n))
    else:
        txt = "Найдите значение выражения $$\\frac{\log_{" + latex(k) + '}' + latex(a) + '}{\log_{' + latex(k) + '}' + latex(int(n))
    if n==10:
        txt = txt + '}+\log' + latex(int(b))
    else:
        txt = txt + '}+\log_{' + latex(int(n)) + '}' + latex(int(b))
    txt = txt.replace('.', '{,}')
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "") + '.$$'
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(otvet), 3):g}'.replace('.', ',')
    i += 1


wb.save("test21.xlsx")