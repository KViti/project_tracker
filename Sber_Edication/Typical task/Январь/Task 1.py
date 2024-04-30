from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2", "Параметр 3"])
number = 100
i=2
amount=3
a = Symbol('a')
b = Symbol('b')
k = Symbol('k')
while (i < number + 2):
    a = random.randint(1, 50)
    b = random.randint(40, 200)
    while ((b + a) ** 2 - 8 * b * a<=0 or ((b+a)-sqrt((b+a)**2-8*b*a))/2<=0):
        a = random.randint(1, 50)
        b = random.randint(40, 200)
        Otvet = ((b+a)+sqrt((b+a)**2-8*b*a))/2
        c = ((b+a)-sqrt((b+a)**2-8*b*a))/2
    if(round(Otvet * 10 ** amount - int(float(Otvet) * 10 ** amount), 5) != 0 or Otvet-a<=0 or 1/b+1/(Otvet-a)-2/Otvet!=0 or Otvet<40):
        continue
    c = random.randint(c, Otvet-1)
    txt = 'Из пункта A в пункт B одновременно выехали два автомобиля. ' \
          + 'Первый проехал с постоянной скоростью весь путь. ' + \
          'Второй проехал первую половину пути со скоростью, меньшей скорости первого на ' \
          + str(a) + ' км/ч, а вторую половину пути — со скоростью ' \
          + str(b) + ' км/ч, в результате чего прибыл в пункт В одновременно с первым автомобилем. Найдите скорость первого автомобиля, если известно, что она больше ' \
          + str(c) + ' км/ч. Ответ дайте в км/ч.'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{b}'
    ws.cell(row=i, column=8).value = f'{c}'
    i+=1

wb.save("test1.xlsx")