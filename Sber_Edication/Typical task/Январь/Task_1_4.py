from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2", "Параметр 3"])
number = 100
i=2
amount=1
a = Symbol('a')
b = Symbol('b')
c = Symbol('c')
while (i < number + 2):
    a = random.randint(10, 30)*10
    b = random.randint(1, 6)*0.5
    c = random.randint(30, 100)
    Otvet = math.ceil((a*c)/(b*1000))
    while ((a*c)/(b*1000)==int((a*c)/(b*1000)) and Otvet>1):
        a = random.randint(10, 30) * 10
        b = random.randint(1, 5) * 0.5
        c = random.randint(30, 100)
        Otvet = math.ceil((a*c)/(b*1000))
    if (b%1==0):
        b=int(b)
    txt = "Для покраски 1 м2 потолка требуется "\
          + str(a)+" г краски. Краска продается в банках по "\
          +str(b).replace('.', ',')+" кг. Сколько банок краски нужно купить для покраски потолка площадью "\
          +str(c)+" м2?"
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{Otvet}'
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{b}'.replace('.', ',')
    ws.cell(row=i, column=8).value = f'{c}'
    i+=1

wb.save("test_1_4.xlsx")