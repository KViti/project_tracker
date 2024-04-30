from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

name_transport = ["проезда в маршрутном такси", "проезда в маршрутке", "проезда в автобусе"]
name_where = ["в этом маршрутном такси", "в этом маршрутке", "в этом автобусе", "в этот театре",
              "на это кино", "на эту выставку", "в этот музей"]
name_relax = ["билета в театр", "билета в кино", "билета на выставку", "билета в музей"]
word_1=["поездок", "посещений"]
word_2=["проезда", "посещения"]
word_3=["снизится", "повысится"]

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2"])
number = 100
i=2
amount=0
len_transport=len(name_transport)
len_relax=len(name_relax)
a = Symbol('a')
b = Symbol('b')
# c = Symbol('c')
while (i < number + 2):
    cost = [random.randint(5, 12) * 10, random.randint(2, 10) * 10, random.randint(1, 7) * 10,
            random.randint(50, 120) * 10, random.randint(30, 100) * 10, random.randint(20, 50) * 10,
            random.randint(5, 12) * 10]
    percent = random.randint(1,10)*5
    Otvet = random.randint(2, 50)
    k = i % (len_transport + len_relax)
    while round(cost[k]*percent/100 * 10 ** amount - int(float(cost[k]*percent/100) * 10 ** amount), 5) != 0:
        percent = random.randint(5, 50)
    ostatok=random.randint(1, int(cost[k]-cost[k]*percent/100))
    if k<len_transport:
        my = name_transport[k]
    else:
        my = name_relax[k-len_transport]
    if (i //10)%2==0:
        a=Otvet*(cost[k]-cost[k]*percent/100)+ostatok
    else:
        a = Otvet * (cost[k] + cost[k] * percent / 100) + ostatok
    txt = "Стоимость " + my + " составляет " + str(cost[k]) + " руб. Какое наибольшее число " \
          + word_1[0 if (k < len_transport) else 1] + ' можно будет совершить '+ name_where[k] +' на '+ str(int(a))+' руб., если цена ' \
          + word_2[0 if (k < len_transport) else 1] + " " + word_3[(i //10)%2] + " на " + str(percent) + "%?"


    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{Otvet}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{cost[k]}'
    ws.cell(row=i, column=7).value = f'{int(a)}'
    i+=1

wb.save("test_1_3.xlsx")