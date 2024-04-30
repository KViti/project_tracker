from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook
def speed(i):#Судно скорость
    if i%4==1:# Моторная лодка
        a=random.randint(50, 200)
    elif i % 4 == 2:# Теплоход
        a=random.randint(25, 80)
    elif i % 4 == 3:# Баржа
        a=random.randint(10, 40)
    elif i % 4 == 0:# Яхта
        a=random.randint(30, 100)
    return a

def transpor_1(i, a) -> str:#Судно
    if i%4==1:
        message = 'Моторная лодка прошла против течения реки ' \
              + str(a) + ' км и вернулась в пункт отправления, затратив на обратный путь на '
    elif i % 4 == 2:
        message = 'Теплоход прошёл против течения реки ' \
                  + str(a) + ' км и вернулся в пункт отправления, затратив на обратный путь на '
    elif i % 4 == 3:
        message = 'Баржа прошла против течения реки ' \
                  + str(a) + ' км и вернулась в пункт отправления, затратив на обратный путь на '
    elif i % 4 == 0:
        message = 'Яхта прошла против течения реки ' \
                  + str(a) + ' км и вернулась в пункт отправления, затратив на обратный путь на '
    return message
def transpor_2(i) -> str:#Судно
    if i%4==1:
        message='меньше. Найдите скорость течения, если скорость лодки в неподвижной воде равна '
    elif i % 4 == 2:
        message = 'меньше. Найдите скорость течения, если скорость теплохода в неподвижной воде равна '
    elif i % 4 == 3:
        message = 'меньше. Найдите скорость течения, если скорость баржи в неподвижной воде равна '
    elif i % 4 == 0:
        message = 'меньше. Найдите скорость течения, если скорость яхты в неподвижной воде равна '
    return message


wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2", "Параметр 3", "Параметр 4"])
number = 100
i=2
amount=1
a = Symbol('a')
# b = Symbol('b')
c = Symbol('c')
while (i < number + 2):
    a = random.randint(10, 500)
    c = speed(i)
    Otvet = random.randint(1, 10)
    while (c-Otvet<=0 or a/(c-Otvet)-a/(c+Otvet)<=0):
        a = random.randint(10, 500)
        c = speed(i)
        Otvet = random.randint(1, 10)
    b = int(a/(c-Otvet)-a/(c+Otvet))
    minutka=(a/(c-Otvet)-a/(c+Otvet)-b)*60
    if(round(b * 10 ** amount - int(b * 10 ** amount), 5) != 0 or round(minutka - int(minutka), 5) != 0):
        continue

    txt = transpor_1(i, a)

    if b==0:
        int(b)
    elif b%10==1 and (b%100<10 or b%100>20):
        txt += str(b) + ' час '
    elif b%10<=4 and b%10!=0 and (b%100<10 or b%100>20):
        txt += str(b) + ' часа '
    else:
        txt += str(b) + ' часов '
    minutka=int(minutka)
    if minutka==0:
        minutka
    elif minutka%10==1 and (minutka%100<10 or minutka%100>20):
        txt += str(int(minutka)) + ' минуту '
    elif minutka%10<=4 and minutka%10!=0 and (minutka%100<10 or minutka%100>20):
        txt += str(int(minutka)) + ' минуты '
    else:
        txt += str(int(minutka)) + ' минут '

    txt += transpor_2(i) + str(c) + ' км/ч. Ответ дайте в км/ч.'

    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{a}'
    ws.cell(row=i, column=7).value = f'{int(b)}'
    ws.cell(row=i, column=8).value = f'{minutka}'
    ws.cell(row=i, column=9).value = f'{c}'
    i+=1

wb.save("test2.xlsx")