from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

name_product = ["Шоколадка", "Пицца", "Упаковка печенья", "Булочка с корицей",
                "Кекс", "Пирожок", "Сладкий батончик", "Рожок мороженого"
                ]
number_product = ["две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять"
                  ]
#Женский род:
#Мужской род:
product_sex = ["две", "два", "одну", "один"]
name_product_234 = ["шоколадки", "пиццы", "упаковки печенья", "булочки с корицей",
                    "кекса", "пирожка", "сладких батончика", "рожка мороженого"
                    ]
name_product_many = ["шоколадок", "пицц", "упаковок печенья", "булочек с корицей",
                     "кексов", "пирожков", "сладких батончиков", "рожков мороженого"
                     ]
name_cost = ["рубль", "рубля", "рублей"]

name_product_56789 = name_product_many

product_cost = [random.randint(30, 120), random.randint(300, 1000), random.randint(100, 500), random.randint(25, 70),
                random.randint(25, 60), random.randint(25, 50), random.randint(35, 80), random.randint(75, 200)]


wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2"])
number = 100
i=2
amount=1
len_product=len(name_product)
a = Symbol('a')
b = Symbol('b')
# c = Symbol('c')
while (i < number + 2):
    # k=random.randint(0, len_product - 1)
    k=i%(len_product - 1)
    my_name_product = name_product[k]
    my_cost = product_cost[k]
    n = 2 + (i - 2) // 20
    m = random.randint(1, 5)
    c = random.randint(1, n - 1)
    ostatok = random.randint(1, my_cost - 1)
    Otvet = (n + 1) * m + c
    a = (n * m + c) * my_cost + ostatok
    txt = my_name_product + " стоит " + str(my_cost) + " "+name_cost[0 if (my_cost % 10 ==1) else (1 if my_cost % 10 !=0 and my_cost % 10 <5 and my_cost % 100 > 20 else 2)]+". В воскресенье в супермаркете действует специальное предложение: заплатив за "

    if n==2:
        txt += product_sex[0 if k<=3 else 1] + " " + name_product_234[k] +", покупатель получает " + number_product[n-1] + " ("+product_sex[2 if k<=3 else 3]+" в подарок). Какое наибольшее количество "\
              + name_product_many[k] +" можно получить, потратив не более "+str(a)+" рублей в воскресенье?"
    elif n<=4:
        txt += number_product[n-2] + " " + name_product_234[k] +", покупатель получает " + number_product[n-1] + " ("+product_sex[2 if k<=3 else 3]+" в подарок). Какое наибольшее количество "\
              + name_product_many[k] +" можно получить, потратив не более "+str(a)+" рублей в воскресенье?"
    elif n<=9:
        txt += number_product[n - 2] + " " + name_product_many[k] + ", покупатель получает " + number_product[n - 1] + " ("+product_sex[2 if k<=3 else 3]+" в подарок). Какое наибольшее количество "\
              + name_product_many[k] + " можно получить, потратив не более " + str(a) + " рублей в воскресенье?"

    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    print(txt)
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{my_cost}'
    ws.cell(row=i, column=7).value = f'{int(a)}'
    i+=1

wb.save("test_1_2.xlsx")