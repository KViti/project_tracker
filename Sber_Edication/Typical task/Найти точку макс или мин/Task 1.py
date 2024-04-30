import random
import openpyxl
from math import sqrt

from sympy import *
import numpy as np
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])

amount=2
number=100
x = Symbol('x', real=True)
a = Symbol('a', real=True)
b = Symbol('b', real=True)
c = Symbol('c', real=True)
i=2
while (i < number + 2):
    if i%2==0:
        a = random.randint(1, 50)
    else:
        a = (-1) * random.randint(1, 50)
    b, c = (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50)
    K = a * x ** 2 + b * x + c
    y = sqrt(K)
    solution = - b / (2 * a)
    while ((a * solution ** 2 + b * solution + c)<=0 or round(solution * 100 - int(solution * 100), 5) != 0 or b==1):  # Пока после запятой более 1 цифры
        if i % 2 == 0:
            a = random.randint(2, 50)
        else:
            a = (-1) * random.randint(2, 50)

        b, c = (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50)
        K = a * x ** 2 + b * x + c
        y = sqrt(K)
        solution = - b / (2 * a)

    if a>0:
        txt = 'Найдите точку минимума функции $$' + latex(y) + '.$$'
    else:
        txt = 'Найдите точку максимума функции $$' + latex(y) + '.$$'

    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(solution), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(solution), 3):g}'.replace('.', ',')
    i+=1

wb.save("test1.xlsx")


















# book = openpyxl.Workbook()
# sheet = book.active
# sheet['B1'].value = 0
#
#
# def TRU_3(a, b, c):  # Функция для генерации нового набора значений
#     a, b, c = random.randint(2, 100), (-1) ** random.randint(1, 2) * random.randint(2, 100), (-1) ** random.randint(1, 2) * random.randint(1, 100)  # Генерация набора значений
#     return a, b, c
#
# def TRU_2(b, c):  # Функция для генерации нового набора значений
#     b, c = (-1) ** random.randint(1, 2) * random.randint(2, 100), (-1) ** random.randint(1, 2) * random.randint(1, 100)  # Генерация набора значений
#     return b, c
#
#
# def znak(t):  # Функция для определения знака
#     if c > 0:
#         return '+'
#     else:
#         return '-'
#
# for i in range(1, 200, 4):
#     a, b, c = random.randint(2, 100), (-1) ** random.randint(1, 2) * random.randint(2, 100), (-1) ** random.randint(1, 2) * random.randint(1, 100)
#
#
#     K = b / (2 * a)  # Ответ
#     while ((c+b*(b/(2*a))-a*(b/(2*a))**2)<0 or round(K * 100 - int(K * 100), 5) != 0 or b==1):  # Пока после запятой более 1 цифры
#         a, b, c = TRU_3(a, b, c)  # создаём новый набор
#         K = b / (2 * a)  # Ответ выражения меняем на новое
#     string = 'Найдите точку максимума функции $$\sqrt{' + str(c) + znak(b) + str(abs(b)) + 'x-' + str(a)+ 'x^2}.$$'
#     sheet[i][0].value = string
#     sheet[i][1].value = K
#
#
#     K = - b / (2 * a)  # Ответ
#     while ((c-b*(b/(2*a))+a*(b/(2*a))**2)<0 or round(K * 100 - int(K * 100), 5) != 0 or b==1):  # Пока после запятой более 1 цифры
#         a, b, c = TRU_3(a, b, c)  # создаём новый набор
#         K = - b / (2 * a)  # Ответ выражения меняем на новое
#     string = 'Найдите точку минимума функции $$\sqrt{' + str(a)+ 'x^2' + znak(b) + str(abs(b)) + 'x' + znak(c) + str(abs(c)) + '}.$$'
#     sheet[i+1][0].value = string
#     sheet[i+1][1].value = K
#
#
#     a = 1
#     K = b / 2  # Ответ
#     while ((c+b*(b/2)-(b/2)**2)<0 or round(K * 100 - int(K * 100), 5) != 0 or b==1):  # Пока после запятой более 1 цифры
#         b, c = TRU_2(b, c)
#         K = b / 2  # Ответ выражения меняем на новое
#     string = 'Найдите точку максимума функции $$\sqrt{'+ str(c)+ znak(b)+ str(abs(b))+ 'x-'+ 'x^2}.$$'
#     sheet[i+2][0].value = string
#     sheet[i+2][1].value = K
#
#
#     K = - b / 2  # Ответ
#     while ((c-b*(b/2)+(b/2)**2)<0 or round(K * 100 - int(K * 100), 5) != 0 or b==1):  # Пока после запятой более 1 цифры
#         b, c = TRU_2(b, c)
#         K = - b / 2  # Ответ выражения меняем на новое
#     string = 'Найдите точку минимума функции $$\sqrt{x^2' + znak(b) + str(abs(b)) + 'x' + znak(c) + str(abs(c)) + '}.$$'
#     sheet[i+3][0].value = string
#     sheet[i+3][1].value = K
#
# book.save("test1.xlsx")
# book.close()
















# print('Пример 1, Максимум:\sqrt[n]{c+bx+ax^2}')
# K = -b / (2 * -a)  # Ответ
# while (round(K * 100 - int(K * 100), 5) != 0):  # Пока после запятой более 1 цифры
#     a, b = TRU(a, b)  # создаём новый набор
#     K = -b / (2 * -a)  # Ответ выражения меняем на новое
#
# # print('a, b, c=', -a, b, c)  # Вид
#
# if (b > 0):
#     # print('Вид выражения: ', u'\u221a', c, '+', b, 'x-', a, 'x^2', sep="")  # Вид
#     print('Найдите точку максимума функции $$\sqrt{', c, '+', b, 'x-', a, 'x^2' '}.$$', sep="")  # LATEX
#     # print('Ответ: -b/(2*-a)=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
# else:
#     # print('Вид выражения: ', u'\u221a', c, '-', abs(b), 'x-', a, 'x^2', sep="")  # Вид
#     print('Найдите точку максимума функции $$\sqrt{', c, '-', abs(b), 'x-', a, 'x^2' '}.$$', sep="")  # LATEX
#     # print('Ответ: -b/(2*-a)=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#
#
# print('Пример 2, Минимум:\sqrt[n]{c+bx+ax^2}')
# K = -b / (2 * a)  # Ответ
# while (round(K * 100 - int(K * 100), 5) != 0):  # Пока после запятой более 1 цифры
#     a, b = TRU(a, b)  # создаём новый набор
#     K = -b / (2 * a)  # Ответ выражения меняем на новое
#
# # print('a, b, c=', a, b, c)  # Вид
#
# if (b > 0):
#     # print('Вид выражения: ', u'\u221a', c, '+', b, 'x+', a, 'x^2', sep="")  # Вид
#     print('Найдите точку минимума функции $$\sqrt{', a ,'x^2', '+', abs(b), 'x', znak, abs(c), '}.$$', sep="")  # LATEX
#     # print('Ответ: -b/(2*-a)=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
# else:
#     # print('Вид выражения: ', u'\u221a', c, '-', abs(b), 'x+', a, 'x^2', sep="")  # Вид
#     print('Найдите точку минимума функции $$\sqrt{', a, 'x^2', '-', abs(b), 'x', znak, abs(c), '}.$$',sep="")  # LATEX
#     # print('Ответ: -b/(2*a)=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#
# print('Пример 3, Максимум:\sqrt[n]{c+bx+x^2}', 'a=-1')
# a = 1
# K = b / 2  # Ответ
# while (round(K * 100 - int(K * 100), 5) != 0):  # Пока после запятой более 1 цифры
#     b = (-1) ^ random.randint(0, 1) * random.randint(2, 100)  # создаём новый набор
#     K = b / 2  # Ответ выражения меняем на новое
#
# # print('a, b, c=', -a, b, c)  # Вид
#
# if (b > 0):
#     # print('Вид выражения: ', u'\u221a', c, '+', b, 'x-', 'x^2', sep="")  # Вид
#     print('Найдите точку максимума функции $$\sqrt{', c, '+', b, 'x-', 'x^2' '}.$$', sep="")  # LATEX
#     # print('Ответ: b/2=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
# else:
#     # print('Вид выражения: ', u'\u221a', c, '-', abs(b), 'x-', 'x^2', sep="")  # Вид
#     print('Найдите точку максимума функции $$\sqrt{', c, '-', abs(b), 'x-', 'x^2' '}.$$', sep="")  # LATEX
#     # print('Ответ: b/2=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#
# print('Пример 4, Минимум:\sqrt[n]{c+bx+x^2}', 'a=1')
# K = -b / 2  # Ответ
# while (round(K * 100 - int(K * 100), 5) != 0):  # Пока после запятой более 1 цифры
#     b = (-1) ^ random.randint(0, 1) * random.randint(2, 100)  # создаём новый набор
#     K = -b / 2  # Ответ выражения меняем на новое
#
# # print('a, b, c=', a, b, c)  # Вид
#
# if (b > 0):
#     # print('Вид выражения: ', u'\u221a', c, '+', b, 'x-', 'x^2', sep="")  # Вид
#     print('Найдите точку минимума функции $$\sqrt{x^2', '+', b, 'x', znak, abs(c), '}.$$', sep="")  # LATEX
#     # print('Ответ: -b/2=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
# else:
#     # print('Вид выражения: ', u'\u221a', c, '-', abs(b), 'x-', 'x^2', sep="")  # Вид
#     print('Найдите точку минимума функции $$\sqrt{x^2', '-', abs(b), 'x', znak, abs(c), '}.$$', sep="")  # LATEX
#     # print('Ответ: -b/2=', str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     print(str(K).rstrip("0").rstrip("."), sep="", end='\n\n')  # значение выражения
#     sheet[i + 3][0].value = string
#     sheet[i + 3][1].value = K
