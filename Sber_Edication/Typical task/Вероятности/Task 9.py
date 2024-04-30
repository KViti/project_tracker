from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Требуемое количество", "Вопрос",
           "Пояснение к вопросу", "Правильно", "Параметр 1", "Параметр 2", "Параметр 3", "Параметр 4"])
number = 100
i=2
amount=3
a = Symbol('a')
b = Symbol('b')
while (i < number + 2):
    sum = random.randint(10, 50)
    a = sum - random.randint(1, sum - 1)
    b = random.randint(1, 5)
    c = b+1
    Otvet = (a/sum)**(b-1)*(1+a/sum)*(sum-a)/sum
    if(round(Otvet * 10 ** amount - int(float(Otvet) * 10 ** amount), 5) != 0):
        continue
    print(Otvet)
    txt = 'Маша коллекционирует принцесс из Киндер-сюрпризов. Всего в коллекции ' + str(sum)\
         + ' разных принцесс, и они равномерно распределены, то есть в каждом очередном Киндер-сюрпризе может с равными вероятностями оказаться любая из ' + str(sum)\
         + ' принцесс. У Маши уже есть '+ str(a) + ' разные принцессы из коллекции. Какова вероятность того, что для получения следующей принцессы Маше придётся купить ещё '+ str(b)\
         + ' или ' + str(c) + ' шоколадных яйца?'
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=3).value = txt
    ws.cell(row=i, column=5).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    ws.cell(row=i, column=6).value = f'{sum}'
    ws.cell(row=i, column=7).value = f'{a}'
    ws.cell(row=i, column=8).value = f'{b}'
    ws.cell(row=i, column=9).value = f'{c}'
    i+=1

wb.save("test9.xlsx")