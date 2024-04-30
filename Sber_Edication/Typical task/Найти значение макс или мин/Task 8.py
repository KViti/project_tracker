from sympy import *
import numpy as np
import random
import math
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.append(["Ссылка на задание", "Текст к заданию (если есть)", "Требуемое количество успешных прохождений", "Вопрос",
           "Пояснение к вопросу", "Правильно"])

amount=2
number=100
x = Symbol('x', real=True)
k = Symbol('k', real=True)
n = Symbol('n', real=True)
k=1
i=2
while (i < number + 1):
    a, b, c = random.randint(2, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50)
    y = a * x ** 2 + b * x + c
    n=random.randint(2, 10)
    eq = log(1/y, n)
    solution = -b / (2 * a)
    K = c-b*(b/(2*a))+a*(b/(2*a))**2 # Ответ

    Otvet = log(K, 1/n)
    if (K<0 or Otvet is zoo or Otvet is nan  or round(Otvet * 10 ** amount - int(float(Otvet) * 10 ** amount), 5) != 0):
        continue
    print(-b/(2*a))
    print(log(K, 1/n))
    max_point_1 = int(solution) - random.randint(3, 10)
    max_point_2 = int(solution) + random.randint(3, 10)
    txt = "Найдите наибольшее значение функции $$\log_\\frac{" + latex(1) + '}{' +latex(n)+'}' + '{(' +latex(y)+')}' + " На отрезке [" + latex(simplify(max_point_1)) + '; ' + latex(simplify(max_point_2)) + "]$$"
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(Otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    i+=2

i = 3
while (i < number + 1):
    a, b, c = (-1) * random.randint(2, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50), (-1) ** random.randint(1, 2) * random.randint(1, 50)
    y = a * x ** 2 + b * x + c
    n=random.randint(2, 10)
    eq = log(1/y, n)
    solution = -b / (2 * a)
    K = c-b*(b/(2*a))+a*(b/(2*a))**2 # Ответ

    Otvet = log(K, 1/n)
    if (K<0 or Otvet is zoo or Otvet is nan  or round(Otvet * 10 ** amount - int(float(Otvet) * 10 ** amount), 5) != 0):
        continue
    print(-b/(2*a))
    print(log(K, 1/n))

    max_point_1 = int(solution) - random.randint(3, 10)
    max_point_2 = int(solution) + random.randint(3, 10)
    txt = "Найдите наименьшее значение функции $$\log_\\frac{" + latex(1) + '}{' +latex(n)+'}' + '{(' +latex(y)+')}' + " На отрезке: [" + latex(simplify(max_point_1)) + '; ' + latex(simplify(max_point_2)) + "]$$"
    txt = txt.replace("\\left(", "")
    txt = txt.replace("\\right)", "")
    ws.cell(row=i, column=4).value = txt
    ws.cell(row=i, column=6).value = f'{round(float(Otvet), 3):g}'
    ws.cell(row=i, column=7).value = f'{round(float(Otvet), 3):g}'.replace('.', ',')
    i+=2

wb.save("test8.xlsx")